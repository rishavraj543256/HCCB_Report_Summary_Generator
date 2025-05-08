import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog
import os
import sys
import logging
import traceback

from copy_template import create_sheet_with_template_if_not_exists

# Logger setup
logger = logging.getLogger("AppLogger")
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler("app.log", encoding="utf-8")
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
if not logger.hasHandlers():
    logger.addHandler(file_handler)

# Get the correct path for template file when running as executable
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    # Use summary_report_template.xlsx for summary_report_tool.py
    if 'template' in relative_path:
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), "summary_report_template.xlsx")
    return os.path.join(base_path, relative_path)


def select_file():
    """Open a file dialog and return the selected file path."""
    root = Tk()
    root.withdraw()  # Hide the main tkinter window
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm")]
    )
    return file_path


def extract_distributor_data(file_path):
    try:
        if not os.path.exists(file_path):
            print("File not found.")
            return

        # Load the workbook
        wb = load_workbook(file_path)

        # Create or ensure Summary Report sheet exists
        create_sheet_with_template_if_not_exists(
            existing_file_path=file_path,
            template_file_path=resource_path("summary_report_template.xlsx"),
            new_sheet_name="Summary Report"
        )
        # 🧠 Reload the workbook to reflect any new sheet creation
        wb = load_workbook(file_path)
        sheet1 = wb['Sheet1']
        summary_sheet = wb['Summary Report']

        # Get existing report numbers in Summary Report
        existing_report_numbers = set()
        for row in summary_sheet.iter_rows(min_row=2, max_col=1):
            report_num = row[0].value
            if report_num:
                existing_report_numbers.add(str(report_num))

        current_row = 1
        distributor_data = []

        while current_row <= sheet1.max_row:
            if sheet1.cell(row=current_row, column=2).value == "Distributor Audit Report":
                distributor_info = {
                    'Report Number': '',
                    'Zone': '',
                    'Unit': '',
                    'Customer Code': '',
                    'Distributor Name': '',
                    'Date of Audit': '',
                    'Book Stock': 0,
                    'Good Stock': 0,
                    'Damaged Stock': 0,
                    'Physical Stock': 0,
                    'Difference': 0,
                    'Products_Over_75': 0,
                    'Observations': set(),  # Using set to store unique observations
                }

                current_row += 1

                while current_row <= sheet1.max_row:
                    cell_value = sheet1.cell(row=current_row, column=2).value
                

                    if cell_value == "Report Number":
                        distributor_info['Report Number'] = str(sheet1.cell(row=current_row, column=4).value)
                    elif cell_value == "Zone":
                        distributor_info['Zone'] = sheet1.cell(row=current_row, column=4).value
                    elif cell_value == "Unit":
                        distributor_info['Unit'] = sheet1.cell(row=current_row, column=4).value
                    elif cell_value == "Customer Code of the Distributor":
                        distributor_info['Customer Code'] = sheet1.cell(row=current_row, column=4).value
                    elif cell_value == "Name of the Distributor":
                        distributor_info['Distributor Name'] = sheet1.cell(row=current_row, column=4).value
                    elif cell_value == "Date of audit":
                        distributor_info['Date of Audit'] = sheet1.cell(row=current_row, column=11).value
                    elif cell_value == "Product Type":
                        break

                    current_row += 1

                while current_row <= sheet1.max_row:
                    if sheet1.cell(row=current_row, column=2).value == "Grand Total":
                        distributor_info['Book Stock'] = sheet1.cell(row=current_row, column=4).value or 0
                        distributor_info['Good Stock'] = sheet1.cell(row=current_row, column=10).value or 0
                        distributor_info['Damaged Stock'] = sheet1.cell(row=current_row, column=11).value or 0
                        distributor_info['Physical Stock'] = sheet1.cell(row=current_row, column=12).value or 0
                        break
                    elif sheet1.cell(row=current_row, column=14).value:
                        try:
                            difference = abs(float(sheet1.cell(row=current_row, column=14).value))
                            if difference > 75:
                                distributor_info['Products_Over_75'] += 1
                        except (ValueError, TypeError):
                            pass
                    
                    # Extract observations from column "Observations/Comments"
                    observation = sheet1.cell(row=current_row, column=15).value  # Column O (15) contains Observations/Comments
                    if observation and str(observation).strip():
                        distributor_info['Observations'].add(str(observation).strip())
                    
                    current_row += 1

                if distributor_info['Report Number'] not in existing_report_numbers:
                    distributor_data.append(distributor_info)

            current_row += 1

        # Append data to Summary Report
        start_row = summary_sheet.max_row + 1
        black = "000000"
        medium_border = Border(
            left=Side(style='thin', color=black),
            right=Side(style='thin', color=black),
            top=Side(style='thin', color=black),
            bottom=Side(style='thin', color=black)
        )

        for data in distributor_data:
            for col_idx in range(1, 19):
                cell = summary_sheet.cell(row=start_row, column=col_idx)
                cell.border = medium_border
            summary_sheet.cell(row=start_row, column=1).value = data['Report Number']
            summary_sheet.cell(row=start_row, column=2).value = data['Zone']
            summary_sheet.cell(row=start_row, column=3).value = data['Unit']
            summary_sheet.cell(row=start_row, column=4).value = data['Customer Code']
            summary_sheet.cell(row=start_row, column=5).value = data['Distributor Name']
            summary_sheet.cell(row=start_row, column=6).value = data['Date of Audit']
            summary_sheet.cell(row=start_row, column=7).value = data['Book Stock']
            summary_sheet.cell(row=start_row, column=8).value = data['Good Stock']
            summary_sheet.cell(row=start_row, column=9).value = data['Damaged Stock']
            summary_sheet.cell(row=start_row, column=10).value = data['Physical Stock']  # Add this line
            
            # Apply formulas: H + I = J (Good Stock + Damaged Stock = Physical Stock)
            good_stock = data['Good Stock'] or 0
            damaged_stock = data['Damaged Stock'] or 0
            summary_sheet.cell(row=start_row, column=10).value = good_stock + damaged_stock
            
            # Apply formulas: G - J = K (Book Stock - Physical Stock = Difference)
            book_stock = data['Book Stock'] or 0
            physical_stock = good_stock + damaged_stock
            summary_sheet.cell(row=start_row, column=11).value = book_stock - physical_stock
            
            summary_sheet.cell(row=start_row, column=12).value = data['Products_Over_75']
            
            # Apply formulas: K - M = O (Difference - Products_Over_75 = Unexplained differences)
            difference = book_stock - physical_stock  # This is the value in column K (11)
            
            # Get the value from column M (13) - this is currently blank/empty
            column_m_value = summary_sheet.cell(row=start_row, column=13).value or 0
            
            # Calculate O (15) = K (11) - M (13)
            summary_sheet.cell(row=start_row, column=15).value = difference - column_m_value

            fill_color=PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
            summary_sheet.cell(row=start_row, column=13).fill = fill_color
            summary_sheet.cell(row=start_row, column=14).fill = fill_color
            summary_sheet.cell(row=start_row, column=16).value = ""
            summary_sheet.cell(row=start_row, column=17).value = ""
            summary_sheet.cell(row=start_row, column=18).value = ""
            
            # Add observations to column N (14)
            # Remove 'Observations/Comments' (case-insensitive, trimmed) from the set
            filtered_observations = [obs for obs in data['Observations'] if obs.strip().lower() != 'observations/comments']
            observations = ', '.join(filtered_observations)
            summary_sheet.cell(row=start_row, column=14).value = observations

            start_row += 1

        # Move the Summary Report sheet to the first position (index 0)
        summary_sheet_index = wb.sheetnames.index("Summary Report")
        wb.move_sheet("Summary Report", offset=-summary_sheet_index)

        # Remove unnecessary ungrouping code (Excel [Group] is a UI feature)
        # (All code related to outline_level, hidden, collapsed, and outlinePr is removed)

        # Preserve any grouping in the workbook
        wb.save(file_path)
        print(f"Summary Report updated successfully with {len(distributor_data)} new records.")
        print("Summary Report sheet moved to the first position.")
    except Exception as e:
        logger.exception(f"Exception in extract_distributor_data: {e}\n{traceback.format_exc()}")
        print(f"Error in extract_distributor_data: {str(e)}")


# === MAIN SCRIPT ===
if __name__ == "__main__":
    try:
        selected_file = select_file()
        if selected_file:
            extract_distributor_data(selected_file)
        else:
            print("No file selected.")
    except Exception as e:
        logger.exception(f"Exception in __main__: {e}\n{traceback.format_exc()}")
