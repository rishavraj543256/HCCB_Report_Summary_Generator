from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_sheet_with_template_if_not_exists(
    existing_file_path,
    template_file_path,
    new_sheet_name="Summary Report",
    position=None
):
    """
    Creates a new sheet with template formatting only if it doesn't exist
    
    Args:
        existing_file_path (str): Path to the existing Excel file
        template_file_path (str): Path to the template Excel file
        new_sheet_name (str): Name for the new sheet
        position (int): Optional position index for the new sheet
        
    Returns:
        tuple: (success: bool, message: str, start_row: int)
    """
    try:
        # Load both workbooks
        template_wb = load_workbook(template_file_path)
        template_ws = template_wb.active
        output_wb = load_workbook(existing_file_path)
        
        # Check if sheet already exists
        if new_sheet_name in output_wb.sheetnames:
            return (False, f"Sheet '{new_sheet_name}' already exists", None)
        
        # Create new sheet at specified position or at end
        if position is not None:
            output_ws = output_wb.create_sheet(title=new_sheet_name, index=position)
        else:
            output_ws = output_wb.create_sheet(title=new_sheet_name)
        
        # Copy all content and formatting from template
        for row in template_ws.iter_rows():
            for cell in row:
                new_cell = output_ws.cell(
                    row=cell.row, 
                    column=cell.column, 
                    value=cell.value
                )
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format

        # Copy merged cells
        for merged_cell_range in template_ws.merged_cells.ranges:
            output_ws.merge_cells(str(merged_cell_range))
        
        # Copy column widths
        for col in template_ws.column_dimensions:
            output_ws.column_dimensions[col].width = \
                template_ws.column_dimensions[col].width
        
        # Copy row heights
        for row in template_ws.row_dimensions:
            output_ws.row_dimensions[row].height = \
                template_ws.row_dimensions[row].height
        
        # Save the workbook
        output_wb.save(existing_file_path)
        
        return (True, f"Created new sheet '{new_sheet_name}' with template format", 1)
        
    except PermissionError:
        return (False, "Error: File is open in another program", None)
    except Exception as e:
        return (False, f"Error: {str(e)}", None)
