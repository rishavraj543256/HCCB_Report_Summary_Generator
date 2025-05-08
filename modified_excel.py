import os
import shutil
import tkinter as tk
from copy import copy
from datetime import datetime
from tkinter import filedialog, messagebox
import logging
import traceback

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter
import sys
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill

# Logger setup
logger = logging.getLogger("AppLogger")
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler("app.log", encoding="utf-8")
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
if not logger.hasHandlers():
    logger.addHandler(file_handler)

def select_file(title="Select File", filetypes=[("Excel files", "*.xlsx")]):
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    return file_path


def select_folder(title="Select Folder"):
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title=title)
    return folder_path


def get_output_preference():
    root = tk.Tk()
    root.withdraw()
    result = messagebox.askquestion(
        "Output Preference",
        "Do you want to generate a new output file?\n\n"
        + "Select 'Yes' for new file\n"
        + "Select 'No' to append to existing file",
    )
    return result == "yes"


# Label text -> (data key, target column letter)
labels_with_column = {
    "Report Number": ("report_number", "D", "D:G"),
    "Zone": ("zone", "D", "D:G"),
    "Unit": ("unit", "D", "D:G"),
    "Customer Code of the Distributor": ("customer_code", "D", "D:G"),
    "Name of the Distributor": ("distributor_name", "D", "D:G"),
    "Address 1": ("address1", "D", "D:G"),
    "Address 2": ("address2", "L", "L:O"),
    "City, State": ("city_state", "L", "L:O"),
    "Audit Firm": ("audit_firm", "L", "L:O"),
    "Audit Team Lead :": ("audit_team_lead", "L", "L:O"),
    "Contact No": ("contact_no", "L", "L:O"),
    "Date of audit": ("date_of_audit", "L", "L:O"),
}



def copy_template():
    template_path = get_resource_path("report_template.xlsx")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_filename = f"template_copy_{timestamp}.xlsx"

    if not os.path.exists(template_path):
        print(f"Template file not found at: {template_path}")
        return None

    try:
        shutil.copy2(template_path, new_filename)

        print(f"Excel file copied successfully to: {new_filename}")
        return new_filename
    except Exception as e:
        print(f"Error copying file: {e}")
        return None
def copy_template_to_existing_file(template_path, existing_output_path):
    template_wb = load_workbook(template_path)
    template_ws = template_wb.active

    output_wb = load_workbook(existing_output_path)
    output_ws = output_wb.active

    # Step 1: Find the last used row
    last_row = output_ws.max_row
    new_start_row = last_row + 4  # 3 blank rows + 1 to start

    # Step 2: Copy row by row
    for row in template_ws.iter_rows():
        target_row_idx = new_start_row + row[0].row - 1
        for cell in row:
            new_cell = output_ws.cell(
                row=target_row_idx, column=cell.column, value=cell.value
            )
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # Step 3: Copy merged cells
    for merged_cell_range in template_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_cell_range.bounds
        shifted_range = f"{get_column_letter(min_col)}{min_row + new_start_row - 1}:{get_column_letter(max_col)}{max_row + new_start_row - 1}"
        output_ws.merge_cells(shifted_range)

    # Step 4: Copy column widths
    for col in template_ws.column_dimensions:
        output_ws.column_dimensions[col].width = template_ws.column_dimensions[
            col
        ].width

    # Step 5: Copy row heights
    for row_dim in template_ws.row_dimensions:
        output_ws.row_dimensions[
            new_start_row + row_dim - 1
        ].height = template_ws.row_dimensions[row_dim].height

    output_wb.save(existing_output_path)
    # apply_thick_border_to_sections(existing_output_path)
    print(f"✅ Template format copied to {existing_output_path} after row {last_row}")

    return new_start_row


# Add this function to sort the categories in the desired order
def sort_categories(grouped_data):
    # Define the desired order of categories
    category_order = ['CSD', 'Maaza', 'Juice', 'Soda', 'Water', 'Milk', 'Others']
    
    # Create a dictionary to store the sorted groups
    sorted_groups = {}
    
    # Add groups in the specified order
    for category in category_order:
        if category in grouped_data.groups:
            sorted_groups[category] = grouped_data.get_group(category)
    
    # Add any remaining groups that weren't in our specified order
    for category in grouped_data.groups:
        if category not in sorted_groups:
            sorted_groups[category] = grouped_data.get_group(category)
    
    return sorted_groups


def find_and_fill(ws, label, column_letter, value, merge_range=None, start_row=1):
    """Find row by label and fill value in fixed column and optionally merge."""
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == label.lower():
                row_number = cell.row

                if merge_range:
                    start_col, end_col = merge_range.split(":")
                    start_col_idx = column_index_from_string(start_col)
                    end_col_idx = column_index_from_string(end_col)
                    cell_range = f"{start_col}{row_number}:{end_col}{row_number}"

                    if cell_range in ws.merged_cells.ranges:
                        ws.unmerge_cells(cell_range)

                    ws.merge_cells(cell_range)
                    ws.cell(row=row_number, column=start_col_idx).value = value
                    ws.cell(row=row_number, column=start_col_idx).alignment = Alignment(
                        vertical="center", horizontal="left"
                    )
                else:
                    ws[f"{column_letter}{row_number}"].value = value

                return


def fill_header_section(ws, matched_row, report_number, start_row=1):
    """Fill headers dynamically based on label names"""
    customer_name = str(matched_row.get("Name of DB", ""))
    if "-" in customer_name:
        number_part, string_part = customer_name.split("-", 1)
    else:
        number_part = string_part = customer_name

    # Create a dictionary with default values for all fields
    distributor_data = {
        "report_number": report_number,
        "zone": matched_row.get("Zone", ""),
        "unit": matched_row.get("Unit", ""),  # Default empty string if Unit not found
        "customer_code": number_part.strip(),
        "distributor_name": string_part.strip(),
        "address1": matched_row.get("Address 1", ""),
        "address2": matched_row.get("Address 2", ""),
        "city_state": f"{matched_row.get('District', '')}, {matched_row.get('State Name', '')}",
        "audit_firm": "RUTUL SHAH & ASSOCIATES",
        "audit_team_lead": matched_row.get("Audit Team Lead", ""),
        "contact_no": matched_row.get("Contact No", ""),
        "date_of_audit": matched_row.get("Date of Audit", ""),
    }

    for label, (key, column, merge_range) in labels_with_column.items():
        find_and_fill(
            ws, label, column, distributor_data.get(key, ""), merge_range, start_row
        )


def get_excel_files_from_folder(folder_path):
    """Get all Excel files from the specified folder"""
    excel_files = []
    for file in os.listdir(folder_path):
        if file.endswith((".xlsx", ".xls")):
            excel_files.append(os.path.join(folder_path, file))
    return excel_files


def calculate_column_sum(ws, column_letter, start_row, end_row):
    """Calculate sum of a column between specified rows"""
    total = 0
    col_idx = column_index_from_string(column_letter)

    for row in range(start_row, end_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if isinstance(cell_value, (int, float)):
            total += cell_value

    return total


def select_excel_files(dump_folder=None, plan_file=None):
    try:
        # Use provided dump folder or prompt for selection
        df1_folder = dump_folder if dump_folder else select_folder(title="Select Folder containing ExampleFiles")
        if not df1_folder:
            print("No folder selected. Exiting...")
            return

        # Use provided plan file or prompt for selection
        df2_path = plan_file if plan_file else select_file(title="Select Distributor Data Excel")
        if not df2_path:
            print("No Distributor Data file selected. Exiting...")
            return

        # Load the distributor data
        df2 = pd.read_excel(df2_path, "Sheet1")

        # Get all Excel files from the selected folder
        excel_files = get_excel_files_from_folder(df1_folder)
        if not excel_files:
            print("No Excel files found in the selected folder. Exiting...")
            return

        # Get output preference
        create_new = get_output_preference()

        output_file = None
        
        if create_new:
            output_file = copy_template()  # Store in output_file instead of new_file
            print(f"New file created: {output_file}")
            if not output_file:
                return
            wb = load_workbook(output_file)
            ws = wb.active
            current_report_number = 1
            new_start_row = 1
        else:
            output_file = select_file(title="Select Existing Output File")  # Store in output_file
            wb = load_workbook(output_file)
            ws = wb.active

            # Find the highest existing report number
            existing_numbers = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip() == "Report Number":
                        report_num_cell = ws.cell(row=cell.row, column=4)
                        if isinstance(report_num_cell.value, (int, float)):
                            existing_numbers.append(int(report_num_cell.value))
            current_report_number = max(existing_numbers) + 1 if existing_numbers else 1
            print(f"Existing file selected: {output_file}")

        # Process each Excel file in the folder
        for file_index, excel_file in enumerate(excel_files):

            

            try:
                print(
                    f"\nProcessing file {file_index + 1} of {len(excel_files)}: {excel_file}"
                )
                df1 = pd.read_excel(excel_file)

                # Find all unique distributor codes in this file
                unique_codes = df1["Distributor code"].unique()
                print(f"Found distributor codes: {unique_codes}")

                # Track if we found any matching codes
                found_any_match = False

                for code in unique_codes:
                    match = df2[df2["Code Of DB"] == code]

                    if not match.empty:
                        found_any_match = True
                        matched_row = match.iloc[0]

                        print(f"Processing distributor code: {code}")

                        # For all files after the first one OR when appending to existing file,
                        # we need to copy the template format
                        if file_index > 0 or not create_new:
                            # Copy template format for the new report section
                            base_dir = os.path.dirname(os.path.abspath(__file__))
                            template_path = os.path.join(base_dir, "report_template.xlsx")
                            new_start_row = copy_template_to_existing_file(
                                template_path, output_file  # Use output_file here
                            )
                            wb = load_workbook(output_file)  # Use output_file here
                            ws = wb.active

                        # Fill header section with current report number
                        fill_header_section(
                            ws, matched_row, current_report_number, new_start_row
                        )

                        # Get data for this distributor
                        distributor_data = df1[df1["Distributor code"] == code]
                        grouped_data = distributor_data.groupby("Item Type")
                        # Sort the categories in the desired order
                        sorted_groups = sort_categories(grouped_data)
                        data_start_row = new_start_row + len(labels_with_column) + 2

                        # Fill item data
                        for category, items in sorted_groups.items():
                            #import ipdb;ipdb.set_trace()
                            ws.cell(row=data_start_row, column=2).value = category
                            for _, row in items.iterrows():
                                ws.cell(row=data_start_row, column=3).value = row[
                                    "Item Name"
                                ]
                                
                                ws.cell(row=data_start_row, column=4).value = row[
                                    "Item QTY As Per book Stock"
                                ]
                                
                                ws.cell(row=data_start_row, column=10).value = row[
                                    "Total Physical Stock"
                                ]
                                
                                ws.cell(row=data_start_row, column=15).value = row[
                                    "Remarks"
                                ]

                                

                                a_val = float(
                                    ws.cell(row=data_start_row, column=4).value or 0
                                )
                                
                                b_val = float(
                                    ws.cell(row=data_start_row, column=5).value or 0
                                )
                                
                                c_val = float(
                                    ws.cell(row=data_start_row, column=6).value or 0
                                )
                                
                                d_val = float(
                                    ws.cell(row=data_start_row, column=7).value or 0
                                )
                                
                                f_val = float(
                                    ws.cell(row=data_start_row, column=9).value or 0
                                )
                                
                                g_val = float(
                                    ws.cell(row=data_start_row, column=10).value or 0
                                )
                                
                                h_val = float(
                                    ws.cell(row=data_start_row, column=11).value or 0
                                )
                                

                                e_result = a_val + b_val - c_val - d_val
                                ws.cell(row=data_start_row, column=8).value = e_result
                                

                                i_result = f_val + g_val + h_val
                                ws.cell(row=data_start_row, column=13).value = i_result
                                

                                j_result = e_result - i_result
                                ws.cell(row=data_start_row, column=14).value = abs(
                                    j_result
                                )
                                

                                ws.cell(
                                    row=data_start_row, column=8
                                ).number_format = "#,##0"
                                ws.cell(
                                    row=data_start_row, column=13
                                ).number_format = "#,##0"
                                ws.cell(
                                    row=data_start_row, column=14
                                ).number_format = "#,##0"

                                data_start_row += 1


                        # ===== PUT THE BORDER CODE HERE =====
                        data_first_row = new_start_row + len(labels_with_column) + 1  # First data row
                        data_last_row = data_start_row - 1  # Last data row (before Grand Total)
                        grand_total_row = data_start_row
                        black = '000000'  
                        thin_border = Border(
                            left=Side(style='thin', color=black),
                            right=Side(style='thin', color=black),
                            top=Side(style='thin', color=black),
                            bottom=Side(style='thin', color=black)
)

                        # Apply borders to all data cells (columns B to N)
                        for row in ws.iter_rows(min_row=data_first_row, max_row=data_last_row, min_col=2, max_col=15):  # Columns B to N
                            for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center')
                                cell.font = Font(bold=True)
                        # import ipdb;ipdb.set_trace()   


                        # Apply borders to grand total row (columns B, D, H, J, M, N)
                        total_cols = [2,3, 4,5,6,7,8,9,10,11,12,13,14,15]  # B, D, H, J, M, N
                        for col in total_cols:
                            cell = ws.cell(row=grand_total_row, column=col)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            cell.font = Font(bold=True)
                        
                        #Apply background color to 'I' column
                        for row in ws.iter_rows(min_row=data_first_row, max_row=data_last_row+1, min_col=9, max_col=9):
                            for cell in row:
                                cell.fill = PatternFill("solid", fgColor="222B35")





                        print(f"Checking borders for cell B{data_first_row}:", ws[f'B{data_first_row}'].border)
                        wb.save(output_file)
                        print(f"Borders applied from row {data_first_row} to {data_last_row}")

                        # Add grand total for this report
                        ws.cell(row=grand_total_row, column=2).value = "Grand Total"

                        # Calculate sums for this report only using the new function
                        ws.cell(
                            row=grand_total_row, column=4
                        ).value = calculate_column_sum(
                            ws,
                            "D",
                            new_start_row + len(labels_with_column) + 2,
                            data_start_row - 1,
                        )
                        ws.cell(
                            row=grand_total_row, column=8
                        ).value = calculate_column_sum(
                            ws,
                            "H",
                            new_start_row + len(labels_with_column) + 2,
                            data_start_row - 1,
                        )
                        ws.cell(
                            row=grand_total_row, column=10
                        ).value = calculate_column_sum(
                            ws,
                            "J",
                            new_start_row + len(labels_with_column) + 2,
                            data_start_row - 1,
                        )
                        ws.cell(
                            row=grand_total_row, column=13
                        ).value = calculate_column_sum(
                            ws,
                            "M",
                            new_start_row + len(labels_with_column) + 2,
                            data_start_row - 1,
                        )
                        ws.cell(
                            row=grand_total_row, column=14
                        ).value = calculate_column_sum(
                            ws,
                            "N",
                            new_start_row + len(labels_with_column) + 2,
                            data_start_row - 1,
                        )

                        # Format the grand total row
                        for col in ["D", "H", "J", "M", "N"]:
                            ws.cell(
                                row=grand_total_row,
                                column=column_index_from_string(col),
                            ).number_format = "#,##0"

                        
                        # Increment report number and update start row for next report
                        current_report_number += 1
                        new_start_row = grand_total_row + 3  # Space for next report
                        # Save after each distributor
                        wb.save(output_file)
                        print(f"Successfully processed distributor code: {code}")
                    else:
                        print(
                            f"Distributor code {code} not found in target file. Skipping..."
                        )

                if not found_any_match:
                    print(f"No matching distributor codes found in file: {excel_file}")
                else:
                    print(f"Saved progress after processing file: {excel_file}")

            except Exception as e:
                logger.exception(f"Error processing file {excel_file}: {str(e)}\n{traceback.format_exc()}")
                print(f"Error processing file {excel_file}: {str(e)}")
                continue

        print(
            f"\nAll files processed successfully. Output saved to: {output_file}"
        )


        # apply_thick_border_to_sections(output_file)

    except Exception as e:
        logger.exception(f"Exception in select_excel_files: {e}\n{traceback.format_exc()}")
        print(f"Error processing files: {str(e)}")


def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    # Use report_template.xlsx for modified_excel.py
    if 'template' in relative_path:
        return os.path.join(base_path, "report_template.xlsx")
    return os.path.join(base_path, relative_path)

if __name__ == "__main__":
    try:
        select_excel_files()
    except Exception as e:
        logger.exception(f"Exception in __main__: {e}\n{traceback.format_exc()}")

